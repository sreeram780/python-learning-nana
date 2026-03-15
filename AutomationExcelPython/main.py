import openpyxl

inv_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inv_file['Sheet1']

product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row+1):
    supplier_name = product_list.cell(product_row, 4).value

    quantity = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation number of products per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] += 1
    else:
        product_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += quantity * float(price)
    else:
        total_value_per_supplier[supplier_name] = quantity * float(price)

    # logic for products with inventory less than 10
    if quantity < 10:
        products_under_10_inv[int(product_number)] = int(quantity)

    #  add value for total inventory price
    inventory_price.value = quantity * float(price)
    print(f"Adding inventory price {inventory_price.value}")

inv_file.save("inventory_total.xlsx")

# print(product_per_supplier)
# print(total_value_per_supplier)
print(products_under_10_inv)
